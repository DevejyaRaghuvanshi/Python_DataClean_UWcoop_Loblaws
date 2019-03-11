'''
Create an Excel file and add a table to one sheet with the most valuable berry products and their price over time.
Dates are constantly increasing day by day instead of showing only the days at which transactions were made.
'''
#import required libraries
import pandas as pd
from functools import reduce #reduce works like filter and map

#Helper function for appending dataframes to an excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

#Set display size and number of showed columns in the terminal output 
pd.set_option('display.max_columns', None)  
pd.set_option('display.expand_frame_repr', False)
pd.set_option('max_colwidth', -1)

#filename
f='MostValuableSKU.xlsx'

#read in the data from the sheets of the excel file into dataframes
dfs_1=pd.read_excel(f,sheet_name='straw-1',header=0, parse_dates=['DELV_DT'])
dfs_2=pd.read_excel(f,sheet_name='straw-2',header=0, parse_dates=['DELV_DT'])
dfs_3=pd.read_excel(f,sheet_name='straw-3',header=0, parse_dates=['DELV_DT'])
dfba_1=pd.read_excel(f,sheet_name='black-1',header=0, parse_dates=['DELV_DT'])
dfba_2=pd.read_excel(f,sheet_name='black-2',header=0, parse_dates=['DELV_DT'])
dfba_3=pd.read_excel(f,sheet_name='black-3',header=0, parse_dates=['DELV_DT'])
dfbu_1=pd.read_excel(f,sheet_name='blue-1',header=0, parse_dates=['DELV_DT'])
dfbu_2=pd.read_excel(f,sheet_name='blue-2',header=0, parse_dates=['DELV_DT'])
dfbu_3=pd.read_excel(f,sheet_name='blue-3',header=0, parse_dates=['DELV_DT'])
dfr_1=pd.read_excel(f,sheet_name='rasp-1',header=0, parse_dates=['DELV_DT'])
dfr_2=pd.read_excel(f,sheet_name='rasp-2',header=0, parse_dates=['DELV_DT'])
dfr_3=pd.read_excel(f,sheet_name='rasp-3',header=0, parse_dates=['DELV_DT'])
dfc_1=pd.read_excel(f,sheet_name='cherries-1',header=0, parse_dates=['DELV_DT'])
dfc_2=pd.read_excel(f,sheet_name='cherries-2',header=0, parse_dates=['DELV_DT'])
dfc_3=pd.read_excel(f,sheet_name='cherries-3',header=0, parse_dates=['DELV_DT'])


#zip the price and time columns together to make the required dataframes with Time and Price per unit for SKU columns
#This is important because when we merge the dataframes, we will want different column names for different SKUs
dfs1 = pd.DataFrame(list(zip(dfs_1['DELV_DT'], dfs_1['price_per_unit'])), columns=['Time', 'price_per_unit_straw-1'])
dfs2 = pd.DataFrame(list(zip(dfs_2['DELV_DT'], dfs_2['price_per_unit'])), columns=['Time', 'price_per_unit_straw-2'])
dfs3 = pd.DataFrame(list(zip(dfs_3['DELV_DT'], dfs_3['price_per_unit'])), columns=['Time', 'price_per_unit_straw-3'])
dfba1 = pd.DataFrame(list(zip(dfba_1['DELV_DT'], dfba_1['price_per_unit'])), columns=['Time', 'price_per_unit_blackberry-1'])
dfba2 = pd.DataFrame(list(zip(dfba_2['DELV_DT'], dfba_2['price_per_unit'])), columns=['Time', 'price_per_unit_blackberry-2'])
dfba3 = pd.DataFrame(list(zip(dfba_3['DELV_DT'], dfba_3['price_per_unit'])), columns=['Time', 'price_per_unit_blackberry-3'])
dfbu1 = pd.DataFrame(list(zip(dfbu_1['DELV_DT'], dfbu_1['price_per_unit'])), columns=['Time', 'price_per_unit_blueberry-1'])
dfbu2 = pd.DataFrame(list(zip(dfbu_2['DELV_DT'], dfbu_2['price_per_unit'])), columns=['Time', 'price_per_unit_blueberry-2'])
dfbu3 = pd.DataFrame(list(zip(dfbu_3['DELV_DT'], dfbu_3['price_per_unit'])), columns=['Time', 'price_per_unit_blueberry-3'])
dfr1 = pd.DataFrame(list(zip(dfr_1['DELV_DT'], dfr_1['price_per_unit'])), columns=['Time', 'price_per_unit_raspberry-1'])
dfr2 = pd.DataFrame(list(zip(dfr_2['DELV_DT'], dfr_2['price_per_unit'])), columns=['Time', 'price_per_unit_raspberry-2'])
dfr3 = pd.DataFrame(list(zip(dfr_3['DELV_DT'], dfr_3['price_per_unit'])), columns=['Time', 'price_per_unit_raspberry-3'])
dfc1 = pd.DataFrame(list(zip(dfc_1['DELV_DT'], dfc_1['price_per_unit'])), columns=['Time', 'price_per_unit_cherries-1'])
dfc2 = pd.DataFrame(list(zip(dfc_2['DELV_DT'], dfc_2['price_per_unit'])), columns=['Time', 'price_per_unit_cherries-2'])
dfc3 = pd.DataFrame(list(zip(dfc_3['DELV_DT'], dfc_3['price_per_unit'])), columns=['Time', 'price_per_unit_cherries-3'])



#list of dataframes to merge
data_frames = [dfs1,dfs2,dfs3,dfba1,dfba2,dfba3,dfbu1,dfbu2,dfbu3,dfr1,dfr2,dfr3,dfc1,dfc2,dfc3]

#merging the dataframes into one dataframe over the Time column
df_merged = reduce(lambda  left,right: pd.merge(left,right,on=['Time'], how='outer'), data_frames)

#Set Time column as index
df_merged=  df_merged.set_index('Time')

#Upscale the time index to daily frequency and using the mean price if the same SKU is sold on the same date
df_merged = df_merged.resample('D').mean()
df_merged = df_merged.fillna(value=0) #Fill NaN values with 0
print (df_merged.head())

append_df_to_excel('MostValuableSKU-Time-Price.xlsx', df_merged,sheet_name='Sheet-1') #append merged dataframe to the Excel file
