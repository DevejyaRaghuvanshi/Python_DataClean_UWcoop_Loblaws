#Import required libraries
import pandas as pd
import matplotlib.pyplot as plt

#Helper functions 

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
#Sets the display size of the output Dataframe 
pd.set_option('display.max_columns', None)  
pd.set_option('display.expand_frame_repr', False)
pd.set_option('max_colwidth', -1)

#Indicate file name and read the excel file in
f='PandasBerrieCherries/BerrieCherries-New.xlsx'
df=pd.read_excel(f, header=0)

#Initialize a dictionary such that different product types associated to the same product are consistent 
typedict={'rasp':['Rasp','RASP'], 'black':['Black','BLACK'], 'blue':['Blue','BLUE'],
'straw':['FRAIS','STRAW'], 'cherries':['CERISE','CHERRIES']}

def sku_ven(ty):
  '''
  Prints the most valuable SKUs associated to the product given (ty) and their total cost. Prints the 
  most valuable vendors and the total amount of money given to them. 
  Value is only dependant on the amount of money spent on either the product or the vendor.
  SKU is ARTCLE_NUM per RCV_SITE_NUM
  
  Affects: Prints two dataframes
  
  sku_ven: Str -> None
  '''
  
  '''
  Initializes two dataframes with for each of the two types of product using typedict dictionary given above
  '''
      dftype_s=df[df['PO_LINE_ITEM_DESC'].str.contains(typedict[ty][0])]


      dftype_c=df[df['PO_LINE_ITEM_DESC'].str.contains(typedict[ty][1])]
      
# Concatonate the two DFs into one DF
      dft=pd.concat([dftype_c,dftype_s]) 
#Group df by SKU
      df_sku=dft.groupby(['ARTCL_NUM','RCV_SITE_NUM'],sort=False).sum() 
  #Sort df in descending order of total money spent
      df_skuf=df_sku.sort_values(by=['PO_LINE_AMT'],ascending=False) 
      print (df_skuf.head()) #Print most valuable SKU DF
      
#Group df by Vendor Number 
      df_ven=dft.groupby(['PO_VEND_NUM'],sort=False).sum()
  #Sort in descending order of total maney spent
      df_venf=df_ven.sort_values(by=['PO_LINE_AMT'],ascending=False)
#Print head of Most valuable Vendors DF
      print (df_venf.head())

      print (df_skuf.index.values.tolist()[0][0])

      #Define MV_SKU dataframes and export to excel sheets
      #1st most valuable product
      df_skuf_1art=df[df['ARTCL_NUM']==(df_skuf.index.values.tolist()[0][0])]
      df_skuf_1=df_skuf_1art[df_skuf_1art['RCV_SITE_NUM']== (df_skuf.index.values.tolist()[0][1])]

      df_skuf_1 = df_skuf_1.assign(price_per_unit=df_skuf_1['PO_LINE_AMT']/df_skuf_1['PO_LINE_QTY'])
      #Sort records according to ascending dates
      df_skuf_1f=df_skuf_1.sort_values(by=['DELV_DT'],ascending=True)

      print(df_skuf_1f.head())

      #2nd most valuable product
      df_skuf_2art=df[df['ARTCL_NUM']==(df_skuf.index.values.tolist()[1][0])]
      df_skuf_2=df_skuf_2art[df_skuf_2art['RCV_SITE_NUM']== (df_skuf.index.values.tolist()[1][1])]

      df_skuf_2 = df_skuf_2.assign(price_per_unit=df_skuf_2['PO_LINE_AMT']/df_skuf_2['PO_LINE_QTY'])
      #Sort records according to ascending dates
      df_skuf_2f=df_skuf_2.sort_values(by=['DELV_DT'],ascending=True)

      print(df_skuf_2f.head())

      #3rd most valuable product
      df_skuf_3art=df[df['ARTCL_NUM']==(df_skuf.index.values.tolist()[2][0])]
      df_skuf_3=df_skuf_3art[df_skuf_3art['RCV_SITE_NUM']== (df_skuf.index.values.tolist()[2][1])]

      df_skuf_3 = df_skuf_3.assign(price_per_unit=df_skuf_3['PO_LINE_AMT']/df_skuf_3['PO_LINE_QTY'])
      #Sort records according to ascending dates
      df_skuf_3f=df_skuf_3.sort_values(by=['DELV_DT'],ascending=True)

      print(df_skuf_3f.head())


      #initialize sheet names per product
      sheet1=ty+'-1'
      sheet2=ty+'-2'
      sheet3=ty+'-3'
      # Create a Pandas Excel writer using XlsxWriter as the engine.
      # Write each dataframe to a different worksheet.

      append_df_to_excel('MostValuableSKU.xlsx', df_skuf_1f, sheet_name=sheet1, index=False)
      append_df_to_excel('MostValuableSKU.xlsx', df_skuf_2f, sheet_name=sheet2, index=False)
      append_df_to_excel('MostValuableSKU.xlsx', df_skuf_3f, sheet_name=sheet3, index=False)

#Plot Graphs

      plt.plot(df_skuf_1f['DELV_DT'], df_skuf_1f['price_per_unit'])
      plt.xlabel('Date')
      plt.ylabel('Price Per Case (CAD)')
      plt.title('Price per case over time (1st)')
      plt.show()

      plt.plot(df_skuf_2f['DELV_DT'], df_skuf_2f['price_per_unit'])
      plt.xlabel('Date')
      plt.ylabel('Price Per Case (CAD)')
      plt.title('Price per case over time (2nd)')
      plt.show()

      plt.plot(df_skuf_3f['DELV_DT'], df_skuf_3f['price_per_unit'])
      plt.xlabel('Date')
      plt.ylabel('Price Per Case (CAD)')
      plt.title('Price per case over time (3rd)')
      plt.show()



      return 1
      

sku_ven('rasp') #example call





